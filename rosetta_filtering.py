import argparse
import os
import sys
import glob
import traceback
import csv
 
import pyrosetta
from pyrosetta import rosetta
 
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
 
 
XML_TEMPLATE = r"""
<ROSETTASCRIPTS>
  <SCOREFXNS>
    <ScoreFunction name="sfxn" weights="beta_nov16">
      <Reweight scoretype="coordinate_constraint" weight="1" />
      <Reweight scoretype="atom_pair_constraint" weight="1" />
      <Reweight scoretype="dihedral_constraint" weight="1" />
      <Reweight scoretype="angle_constraint" weight="1" />
    </ScoreFunction>
 
    <ScoreFunction name="sfxn_cart" weights="beta_nov16_cart">
      <Reweight scoretype="coordinate_constraint" weight="1" />
      <Reweight scoretype="atom_pair_constraint" weight="1" />
      <Reweight scoretype="dihedral_constraint" weight="1" />
      <Reweight scoretype="angle_constraint" weight="1" />
    </ScoreFunction>
  </SCOREFXNS>
 
  <RESIDUE_SELECTORS>
    <Chain name="chainA" chains="{PROTEIN_CHAIN}"/>
    <Chain name="chainB" chains="{PEPTIDE_CHAIN}"/>
 
    <Neighborhood name="interface_chA" selector="chainB" distance="14.0" />
    <Neighborhood name="interface_chB" selector="chainA" distance="14.0" />
    <And name="AB_interface" selectors="interface_chA,interface_chB" />
    <Not name="Not_interface" selector="AB_interface" />
  </RESIDUE_SELECTORS>
 
  <TASKOPERATIONS>
    <ProteinInterfaceDesign name="pack_long"
      design_chain1="0"
      design_chain2="0"
      jump="1"
      interface_distance_cutoff="15"/>
    <OperateOnResidueSubset name="restrict_to_interface" selector="Not_interface">
      <PreventRepackingRLT/>
    </OperateOnResidueSubset>
  </TASKOPERATIONS>
 
  <MOVERS>
    <PeptideCyclizeMover name="pcm" residue_selector="chainB"/>
 
    <TaskAwareMinMover name="minimize_interface"
      scorefxn="sfxn_cart"
      tolerance="0.01"
      cartesian="true"
      task_operations="restrict_to_interface"
      jump="0" />
 
    <TaskAwareMinMover name="min"
      scorefxn="sfxn"
      bb="0"
      chi="1"
      task_operations="pack_long" />
  </MOVERS>
 
  <FILTERS>
    <Ddg name="ddg"
      threshold="50"
      jump="1"
      repeats="5"
      repack="1"
      relax_mover="min"
      confidence="0"
      scorefxn="sfxn"
      extreme_value_removal="1" />
 
    <ContactMolecularSurface name="contact_molecular_surface"
      distance_weight="0.5"
      target_selector="chainA"
      binder_selector="chainB"
      confidence="0" />
  </FILTERS>
 
  <SIMPLE_METRICS>
    <SapScoreMetric name="sap_score" score_selector="chainB" />
  </SIMPLE_METRICS>
 
  <PROTOCOLS>
    <Add mover="pcm" />
    <Add mover="minimize_interface" />
    <Add mover="pcm" />
    <Add filter="ddg" />
    <Add metrics="sap_score" />
    <Add filter="contact_molecular_surface" />
  </PROTOCOLS>
</ROSETTASCRIPTS>
"""


def peptide_from_filename(pdb_path: str) -> str:
    base = os.path.splitext(os.path.basename(pdb_path))[0]
    # common naming: <peptide>_on_monomer_..._prediction
    if "_on_" in base:
        return base.split("_on_")[0]
    if base.endswith("_prediction"):
        return base[:-11]
    return base


def detect_protein_and_peptide_chain_ids(pose):
    """
    Heuristic: peptide = shortest polymer chain; protein = longest polymer chain.
    Returns (protein_chain_id, peptide_chain_id).
    """
    pdbi = pose.pdb_info()
    chain_map = {}
    for i in range(1, pose.total_residue() + 1):
        ch = pdbi.chain(i)
        chain_map.setdefault(ch, []).append(i)

    chain_poly_lengths = []
    for ch, idxs in chain_map.items():
        poly = [i for i in idxs if pose.residue(i).is_polymer()]
        if poly:
            chain_poly_lengths.append((ch, len(poly)))

    if len(chain_poly_lengths) < 2:
        raise RuntimeError("Could not detect >=2 polymer chains (need protein+peptide).")

    chain_poly_lengths.sort(key=lambda x: x[1])
    peptide_chain = chain_poly_lengths[0][0]
    protein_chain = chain_poly_lengths[-1][0]
    return protein_chain, peptide_chain


def write_csv(rows, csv_path, fieldnames):
    with open(csv_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_xlsx(rows, xlsx_path, fieldnames, sheet_name="results"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # header
    header_font = Font(bold=True)
    ws.append(fieldnames)
    for j in range(1, len(fieldnames) + 1):
        ws.cell(row=1, column=j).font = header_font
    ws.freeze_panes = "A2"

    # rows
    for r in rows:
        ws.append([r.get(k, "") for k in fieldnames])

    # reasonable column widths
    for j, key in enumerate(fieldnames, start=1):
        max_len = len(key)
        for i in range(2, ws.max_row + 1):
            v = ws.cell(row=i, column=j).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(j)].width = min(max_len + 2, 60)

    wb.save(xlsx_path)


def main():
    ap = argparse.ArgumentParser(description="Run RosettaScripts over a folder; output PDBs + CSV + XLSX.")
    ap.add_argument("--input_dir", required=True, help="Folder with input PDB complexes")
    ap.add_argument("--output_dir", required=True, help="Folder for output PDBs")
    ap.add_argument("--csv", default="results.csv", help="Output CSV path")
    ap.add_argument("--xlsx", default="results.xlsx", help="Output XLSX path")
    ap.add_argument("--pattern", default="*.pdb", help="Glob pattern inside input_dir (default: *.pdb)")
    ap.add_argument("--mute", action="store_true", help="Mute Rosetta output")
    args = ap.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    init_opts = " -corrections:beta_nov16 true "
    if args.mute:
        init_opts = " -mute all " + init_opts
    pyrosetta.init(init_opts)

    pdbs = sorted(glob.glob(os.path.join(args.input_dir, args.pattern)))
    if not pdbs:
        print(f"ERROR: no files matched {args.pattern} in {args.input_dir}", file=sys.stderr)
        sys.exit(1)

    rows = []
    fieldnames = [
        "peptide",
        "input_pdb",
        "output_pdb",
        "protein_chain",
        "peptide_chain",
        "ddg",
        "contact_molecular_surface",
        "sap_score",
        "status",
        "error",
    ]

    for pdb in pdbs:
        name = os.path.basename(pdb)
        pep = peptide_from_filename(pdb)
        out_pdb = os.path.join(args.output_dir, name)

        print(f"Processing {name} ...", file=sys.stderr)

        row = {
            "peptide": pep,
            "input_pdb": name,
            "output_pdb": os.path.basename(out_pdb),
            "protein_chain": "",
            "peptide_chain": "",
            "ddg": "",
            "contact_molecular_surface": "",
            "sap_score": "",
            "status": "FAIL",
            "error": "",
        }

        try:
            pose = rosetta.core.import_pose.pose_from_file(pdb)

            protein_chain, peptide_chain = detect_protein_and_peptide_chain_ids(pose)
            row["protein_chain"] = protein_chain
            row["peptide_chain"] = peptide_chain

            xml = XML_TEMPLATE.format(PROTEIN_CHAIN=protein_chain, PEPTIDE_CHAIN=peptide_chain)
            xml_objs = rosetta.protocols.rosetta_scripts.XmlObjects.create_from_string(xml)

            protocol = xml_objs.get_mover("ParsedProtocol")
            ddg_filter = xml_objs.get_filter("ddg")
            cms_filter = xml_objs.get_filter("contact_molecular_surface")
            sap_metric = xml_objs.get_simple_metric("sap_score")

            protocol.apply(pose)

            # Write clean PDB (no metric text appended)
            pose.dump_pdb(out_pdb)

            ddg_val = ddg_filter.report_sm(pose)
            cms_val = cms_filter.report_sm(pose)
            sap_val = sap_metric.calculate(pose)

            row["ddg"] = f"{ddg_val:.6f}"
            row["contact_molecular_surface"] = f"{cms_val:.6f}"
            row["sap_score"] = f"{sap_val:.6f}"
            row["status"] = "OK"
            row["error"] = ""

        except Exception as e:
            row["error"] = str(e)
            # keep going, but record failure
            print(f"FAILED on {name}: {e}", file=sys.stderr)
            # If you want the full stack trace:
            traceback.print_exc(file=sys.stderr)

        rows.append(row)

    # Write tables (all peptides, including FAIL)
    write_csv(rows, args.csv, fieldnames)
    write_xlsx(rows, args.xlsx, fieldnames)

    print("Done.", file=sys.stderr)
    print(f"Output PDBs: {args.output_dir}", file=sys.stderr)
    print(f"Results CSV: {args.csv}", file=sys.stderr)
    print(f"Results XLSX: {args.xlsx}", file=sys.stderr)


if __name__ == "__main__":
    main()
